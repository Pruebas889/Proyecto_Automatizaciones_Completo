# coding: utf-8
import pdfplumber
import os
import re
import sys
import glob
import logging
import time
from datetime import datetime
import PyPDF2

# Configuración de logs
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

if sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)

# Configuración
PDF_DIR = r"C:\Users\jperdomolc\Pictures\Proyecto_Automatizaciones_Completo\Automatizacion_PDF\paginas_separadas"

# 🔥 PARÁMETROS DE LOTE (PAUSAS)
FACTURAS_POR_LOTE = 15  # Procesa 15 facturas por lote
PAUSA_ENTRE_LOTES = 30  # Pausa de 30 segundos entre lotes

HEADERS_BASE = [
    'Fecha_Ejecucion',
    'Empresa',
    'Numero_Cliente',
    'Nombre_Cliente',
    'Cuenta_Padre',
    'Direccion',
    'Fecha_Inicial',
    'Fecha_Final',
    'Fecha_Inicial_Aseo',
    'Fecha_Final_Aseo',
    'Total_Energia',
    'Total_Aseo',
    'Lectura_Actual',
    "Energia_Consumida",
    'Consumo_Unidad',
    'Costo_Unitario',
    'Estado_Pago',
    'NIT_Empresa',
    'NIT_Aseo',
    'Cuenta_Contrato',
    'Prestador_Aseo'
    
]

HEADERS = HEADERS_BASE.copy()

# Archivo de log de seguimiento
LOG_FILE = 'procesamiento_facturas.log'
file_handler = logging.FileHandler(LOG_FILE, encoding='utf-8')
file_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
logging.getLogger().addHandler(file_handler)


def extraer_paginas_pdf(pdf_path):
    """Extrae texto de cada página del PDF por separado"""
    try:
        paginas = []
        with open(pdf_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            total_pages = len(reader.pages)
            
            logging.info(f"📄 Total de páginas en PDF: {total_pages}")
            
            for num_pagina, page in enumerate(reader.pages, 1):
                page_text = page.extract_text()
                if page_text and len(page_text.strip()) > 30:
                    paginas.append({
                        'numero': num_pagina,
                        'texto': page_text
                    })
                else:
                    logging.info(f"⏭️ Página {num_pagina} ignorada (en blanco o sin contenido útil)")
        
        return paginas
    except Exception as e:
        logging.error(f"❌ Error al leer PDF {os.path.basename(pdf_path)}: {e}")
        return []


def limpiar_numero(valor_str):
    """Limpia y convierte número del formato colombiano"""
    if not valor_str:
        return None
    valor_str = str(valor_str).strip()
    try:
        return float(valor_str.replace('.', '').replace(',', '.'))
    except ValueError:
        return None


def extraer_prestador_por_coordenadas(pdf_path, numero_pagina):
    """Extrae el PRESTADOR buscando en zona superior del PDF"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            page = pdf.pages[numero_pagina - 1]

            # 🔥 ZONA SUPERIOR DONDE ESTÁ "ASEO / PRESTADOR"
            bbox = (0, 0, 600, 250)
            texto = page.within_bbox(bbox).extract_text()

            if not texto:
                return None

            texto = texto.replace('\n', ' ')
            texto = re.sub(r'\s+', ' ', texto)

            # 🔥 BUSCAR ENTRE PRESTADOR y NIT
            match = re.search(r'PRESTADOR:\s*(.*?)\s*NIT', texto, re.IGNORECASE)

            if match:
                prestador = match.group(1).strip()
                if len(prestador) > 3:
                    logging.info(f"   → Prestador (coord): {prestador}")
                    return prestador

    except Exception as e:
        logging.warning(f"⚠️ Error coordenadas prestador: {e}")

    return None


def extraer_direccion(texto_pagina, numero_cliente):
    """Extrae SOLO la dirección - Sin el nombre del cliente"""
    try:
        lineas = [l.strip() for l in texto_pagina.split('\n') if l.strip()]

        direccion_lineas = []
        capturar = False

        for i, linea in enumerate(lineas):
            if numero_cliente in linea:
                capturar = True
                continue

            if capturar:
                if re.match(r'^(CL|KR|AC|DG|AV|AVENIDA|CALLE|CARRERA|AK|PZ|TV|VT|RA|AP|MNZ|ZN|BRR|SM|LT|MZ)\s+', linea, re.IGNORECASE):
                    direccion_lineas.append(linea)

                    for siguiente in lineas[i+1:i+5]:
                        if re.search(r'(ruta|transformador|nivel|red|circuito|tipo de servicio|estrato|barrio)', siguiente, re.IGNORECASE):
                            break
                        if not siguiente or siguiente.isdigit():
                            break
                        direccion_lineas.append(siguiente)

                    break

        if direccion_lineas:
            direccion = " ".join(direccion_lineas)
            direccion = re.sub(r'\s+', ' ', direccion).strip()
            return direccion

    except Exception as e:
        logging.warning(f"Error extrayendo dirección: {e}")

    return None

    
def extraer_total_aseo(texto_pagina):
    """Extrae el TOTAL ASEO real desde facturas Enel"""
    if not texto_pagina:
        return 0.0

    texto = texto_pagina.replace('\xa0', ' ')
    texto = re.sub(r'\s+', ' ', texto)

    match_aseo = re.search(r'ASEO\s*-', texto, re.IGNORECASE)

    if not match_aseo:
        return 0.0

    inicio = match_aseo.start()
    texto_aseo = texto[inicio:inicio + 800]

    valores = re.findall(r'\$\s*[\d\.,]+', texto_aseo)

    if not valores:
        return 0.0

    for valor in reversed(valores):
        numero = valor.replace('$', '').replace('.', '').replace(',', '')

        if numero.isdigit():
            numero_int = int(numero)

            if numero_int > 100:
                logging.info(f"   → Total Aseo detectado correctamente: ${numero_int}")
                return float(numero_int)

    return 0.0


def extraer_nit_aseo(texto_pagina):
    """Extrae el NIT de la sección ASEO (robusto)"""
    try:
        texto = texto_pagina.replace('\xa0', ' ')

        match_aseo = re.search(r'ASEO', texto, re.IGNORECASE)
        if not match_aseo:
            return None

        inicio = match_aseo.start()
        bloque = texto[inicio:inicio + 800]

        match_nit = re.search(r'(\d{9,10}-\d)', bloque)
        if match_nit:
            nit = match_nit.group(1)
            logging.info(f"   → NIT ASEO detectado: {nit}")
            return nit

    except Exception as e:
        logging.warning(f"Error extrayendo NIT ASEO: {e}")

    return None


def extraer_cuenta_contrato(texto_pagina, pdf_path=None, numero_pagina=None):
    """Extrae cuenta contrato basada en estructura real ENEL"""
    try:
        texto = texto_pagina.replace('\xa0', ' ')
        lineas = [l.strip() for l in texto.split('\n') if l.strip()]

        for i, linea in enumerate(lineas):

            # 🔥 Detectar NIT (ej: 830131031-1)
            if re.search(r'\d{9,10}-\d', linea):

                # 🔥 siguiente línea = cuenta contrato
                if i + 1 < len(lineas):
                    posible = lineas[i + 1].strip()

                    # 🔥 buscar número dentro de la línea (por si viene mezclado)
                    match_num = re.search(r'\d{6,12}', posible)

                    if match_num:
                        cuenta = match_num.group(0)
                        logging.info(f"   → Cuenta Contrato (debajo NIT): {cuenta}")
                        return cuenta
                        

        logging.info("   ❌ Cuenta Contrato no encontrada por estructura")
        return None

    except Exception as e:
        logging.warning(f"Error extrayendo cuenta contrato: {e}")
        return None


def extraer_periodo_aseo(texto_pagina):
    """Extrae periodo de facturación del bloque ASEO"""
    try:
        texto = texto_pagina.replace('\xa0', ' ')
        lineas = [l.strip() for l in texto.split('\n') if l.strip()]

        for i, linea in enumerate(lineas):

            # 🔥 detectar formato fecha típico
            if re.search(r'\d{2}\.\d{2}\.\d{4}', linea):

                # buscar patrón completo en la misma línea
                match = re.search(
                    r'(\d{2}\.\d{2}\.\d{4}\s*(?:al|-|a)\s*\d{2}\.\d{2}\.\d{4})',
                    linea,
                    re.IGNORECASE
                )

                if match:
                    periodo = match.group(1).strip()

                    # 🔥 FORMATEAR FECHAS (puntos → slash)
                    periodo = periodo.replace('.', '/')

                    logging.info(f"   → Periodo ASEO: {periodo}")
                    return periodo

        logging.info("   ⚠️ Periodo ASEO no encontrado")
        return None

    except Exception as e:
        logging.warning(f"Error extrayendo periodo ASEO: {e}")
        return None

def extraer_prestador_aseo(texto_pagina):
    """Extrae el PRESTADOR buscando empresas conocidas o patrón genérico"""
    try:
        texto = texto_pagina.replace('\xa0', ' ')
        
        match_aseo = re.search(r'ASEO', texto, re.IGNORECASE)
        if not match_aseo:
            return None
        
        inicio = match_aseo.start()
        bloque = texto[inicio:inicio + 2000]
        
        match = re.search(r'([A-Z][A-Z\s\.\-&0-9]+?(?:SAS\s+ESP|LTDA|S\.A\.?|COLOMBIA))', bloque)
        if match:
            prestador = match.group(1).strip()
            prestador = re.sub(r'\s+', ' ', prestador)
            
            if 5 < len(prestador) < 60:
                logging.info(f"   → Prestador ASEO detectado: {prestador}")
                return prestador

    except Exception as e:
        logging.warning(f"Error extrayendo prestador ASEO: {e}")

    return None


def extraer_conceptos_aseo(texto_pagina):
    """
    Extrae TODOS los conceptos de ASEO:
    1. Los viejos predefinidos (con la lógica anterior)
    2. Los nuevos dinámicos (encontrados automáticamente)
    """
    try:
        texto = texto_pagina.replace('\xa0', ' ')
        texto = re.sub(r'\s+', ' ', texto)  # 🔥 NORMALIZAR ESPACIOS
        
        match_aseo = re.search(r'ASEO', texto, re.IGNORECASE)
        if not match_aseo:
            return {}
        
        inicio = match_aseo.start()
        bloque = texto[inicio:inicio + 3000]  # 🔥 AUMENTADO A 3000
        
        resultados = {}
        
        logging.info(f"   🔍 Buscando conceptos ASEO en bloque de {len(bloque)} caracteres...")
        
        # ==================== VIEJOS PREDEFINIDOS ====================
        # 1. SERVICIO NO RESIDENCIAL
        match = re.search(r'SERVICIO\s+NO\s+RESIDENCIAL[^\$]*?\$\s*([\d\.,\-]+)', bloque, re.IGNORECASE)
        if match:
            resultados["Aseo_Servicio_No_Residencial"] = match.group(1).strip()
            logging.info(f"   → Aseo_Servicio_No_Residencial: ${match.group(1).strip()}")
        
        # 2. CONTRIBUCIÓN NO RESIDEN
        match = re.search(r'CONTRIBUCIÓN\s+NO\s+RESIDEN[^\$]*?\$\s*([\d\.,\-]+)', bloque, re.IGNORECASE)
        if match:
            resultados["Aseo_Contribucion_No_Residen"] = match.group(1).strip()
            logging.info(f"   → Aseo_Contribucion_No_Residen: ${match.group(1).strip()}")
        
        # 3. AJUSTE A LA DECENA
        match = re.search(r'AJUSTE\s+A\s+LA\s+DECENA[^\$]*?\$\s*([\d\.,\-]+)', bloque, re.IGNORECASE)
        if match:
            resultados["Aseo_Ajuste_Decena"] = match.group(1).strip()
            logging.info(f"   → Aseo_Ajuste_Decena: ${match.group(1).strip()}")
        
        # 4. RELIQUIDACIÓN
        match = re.search(r'RELIQUIDACIÓN[^\$]*?\$\s*([\d\.,\-]+)', bloque, re.IGNORECASE)
        if match:
            resultados["Aseo_Reliquidacion"] = match.group(1).strip()
            logging.info(f"   → Aseo_Reliquidacion: ${match.group(1).strip()}")
        
        # 5. MENOR VALOR A FACTURA
        match = re.search(r'MENOR\s+VALOR\s+A\s+FACTURA[^\$]*?\$\s*([\d\.,\-]+)', bloque, re.IGNORECASE)
        if match:
            resultados["Aseo_Menor_Valor"] = match.group(1).strip()
            logging.info(f"   → Aseo_Menor_Valor: ${match.group(1).strip()}")
        
        # ==================== NUEVOS DINÁMICOS ====================
        # 🔥 REGEX MÁS FLEXIBLE - Busca "ASEO -" seguido de CUALQUIER COSA hasta $
        matches = re.finditer(r'ASEO\s*-\s*([^$\n]+?)\s*\$\s*([\d\.,\-]+)', bloque, re.IGNORECASE)
        
        matches_encontrados = list(matches)
        logging.info(f"   🔍 Se encontraron {len(matches_encontrados)} conceptos dinámicos ASEO")
        
        for match in matches_encontrados:
            concepto_nombre = match.group(1).strip()
            valor = match.group(2).strip()
            
            # Limpiar el nombre del concepto
            concepto_nombre = re.sub(r'\s+', ' ', concepto_nombre)
            
            # Crear nombre de columna dinámico
            nombre_columna = re.sub(r'[^\w\s]', '', concepto_nombre)
            nombre_columna = nombre_columna.replace(' ', '_').upper()
            nombre_columna = f"ASEO_{nombre_columna}"
            
            # 🔥 NO agregar si ya existe en los predefinidos
            if nombre_columna not in resultados:
                resultados[nombre_columna] = valor
                logging.info(f"   ✅ {nombre_columna}: ${valor}")
            else:
                logging.info(f"   ℹ️ {nombre_columna} ya existe (predefinido)")
        
        logging.info(f"   📊 Total conceptos encontrados: {len(resultados)}")
        return resultados
        
    except Exception as e:
        logging.warning(f"Error extrayendo conceptos ASEO: {e}")
        import traceback
        logging.warning(traceback.format_exc())
        return {}

def extraer_costo_unitario_por_coordenadas(pdf_path, numero_pagina):
    """Extrae el COSTO UNITARIO - BÚSQUEDA DESDE TABLAS DIRECTAMENTE"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            page = pdf.pages[numero_pagina - 1]
            
            # 🔥 ESTRATEGIA 1: Buscar en TABLAS (la más confiable para ENEL)
            logging.info(f"   🔍 Buscando en tablas...")
            
            tablas = page.extract_tables()
            if tablas:
                for num_tabla, tabla in enumerate(tablas):
                    logging.info(f"      📊 Tabla {num_tabla + 1}: {len(tabla)} filas")
                    
                    if not tabla or len(tabla) < 2:
                        continue
                    
                    # 🔥 Buscar encabezados en la primera fila
                    encabezados = tabla[0] if tabla else []
                    indice_valor_unitario = None
                    indice_consumo = None
                    
                    # Buscar índices de columnas
                    for col_idx, celda in enumerate(encabezados):
                        if not celda:
                            continue
                        celda_str = str(celda).strip()
                        
                        if 'Valor' in celda_str and 'Unitario' in celda_str:
                            indice_valor_unitario = col_idx
                            logging.info(f"         ✅ Columna 'Valor Unitario' en índice {col_idx}")
                        
                        if 'CONSUMO' in celda_str.upper():
                            indice_consumo = col_idx
                    
                    # 🔥 Si encontramos la columna "Valor Unitario", buscar el valor
                    if indice_valor_unitario is not None:
                        # Buscar la fila de datos (no encabezado)
                        for num_fila, fila in enumerate(tabla[1:], 1):  # Saltar encabezados
                            # Verificar si es la fila correcta (contiene "CONSUMO")
                            if fila and len(fila) > 0:
                                primera_celda = str(fila[0]).upper() if fila[0] else ""
                                
                                if 'CONSUMO' in primera_celda and 'ACTIVA' in primera_celda:
                                    logging.info(f"      ✅ Encontrada fila CONSUMO ACTIVA SENCILLA")
                                    
                                    # Obtener el valor de la columna "Valor Unitario"
                                    if indice_valor_unitario < len(fila):
                                        valor_celda = fila[indice_valor_unitario]
                                        
                                        if valor_celda:
                                            valor = str(valor_celda).replace('$', '').strip()
                                            logging.info(f"      ✅ Valor Unitario desde tabla: ${valor}")
                                            return valor
            
            # 🔥 ESTRATEGIA 2: Si no está en tablas, buscar en texto
            logging.info(f"   🔍 Fallback: Buscando por patrón en texto...")
            
            texto = page.extract_text()

            # =========================================
            # 🔥 MÉTODO DEFINITIVO (PRIMERA PRIORIDAD)
            # =========================================
            if texto:
                for linea in texto.split('\n'):
                    if 'CONSUMO ACTIVA SENCILLA' in linea.upper():
                        logging.info(f"   📌 Línea consumo encontrada: {linea}")

                        valores = re.findall(r'\$([\d\.,]+)', linea)

                        if len(valores) >= 2:
                            valor_unitario = valores[0]

                            # limpiar formato colombiano
                            valor_limpio = valor_unitario.replace('.', '').replace(',', '.')

                            logging.info(f"   💰 Costo Unitario REAL: {valor_limpio}")

                            return float(valor_limpio)

            # 🔥 NUEVO MÉTODO DEFINITIVO (PRIORIDAD ALTA)
            logging.info(f"   🔍 Buscando en línea CONSUMO ACTIVA SENCILLA...")

            patron = r'CONSUMO\s+ACTIVA\s+SENCILLA.*?\$([\d\.,]+)\s+\$([\d\.,]+)'
            match = re.search(patron, texto, re.IGNORECASE)

            if match:
                costo_unitario = match.group(1).strip()
                logging.info(f"   ✅ Costo Unitario (CONSUMO): ${costo_unitario}")
                return costo_unitario
            
            # Buscar patrón específico: "Valor Unitario" seguido de número
            matches = re.finditer(r'Valor\s+Unitario[:\s]*\$?\s*([\d\.,]+)', texto, re.IGNORECASE)
            valores_encontrados = list(matches)
            
            if valores_encontrados:
                valor = valores_encontrados[0].group(1).strip()
                logging.info(f"   ✅ Valor Unitario por patrón: ${valor}")
                return valor
            
            logging.info(f"   ❌ No se pudo extraer costo unitario")
            return None

    except Exception as e:
        logging.warning(f"Error costo unitario: {e}")
        import traceback
        logging.warning(traceback.format_exc())

    return None


def extraer_datos_factura(texto_pagina, pdf_path=None, numero_pagina=None):
    """Extrae datos de una factura de UNA PÁGINA"""
    datos = {
        "Fecha_Ejecucion": None,
        "Empresa": "ENEL",
        "Numero_Cliente": None,
        "Nombre_Cliente": None,
        "Cuenta_Padre": None,
        "Direccion": None,
        "Periodo_Facturacion": None,
        "Fecha_Inicial": None,
        "Fecha_Final": None,
        "Fecha_Inicial_Aseo": None,
        "Fecha_Final_Aseo": None,
        "Total_Energia": None,
        "Total_Aseo": None,
        "Lectura_Actual": None,
        "Energia_Consumida": None,
        "Consumo_Unidad": "kWh",
        "Costo_Unitario": None,
        "Estado_Pago": None,
        "NIT_Empresa": None,
        "Cuenta_Contrato": None,
        "NIT_Aseo": None,
        "Prestador_Aseo": None
        
    }

    # 🔥 FECHA DE PROCESO (HOY)
    datos["Fecha_Ejecucion"] = datetime.now().strftime("%d/%m/%Y")
    
    # Validar que sea una factura ENEL válida
    # ✅ NUEVA VALIDACIÓN (más flexible y segura)
    texto_normalizado = re.sub(r'\s+', ' ', texto_pagina.upper())
    if "TIPO DE LECTURA" not in texto_normalizado or "REAL" not in texto_normalizado:
        # Solo rechaza si definitivamente no es factura ENEL
        match = re.search(r'(\d{7}-\d)(?:\s*[\d\w])', texto_pagina)  # ⬅️ CAMBIO: [\d\w] en lugar de [A-Z]
        if not match:
            logging.info(f"   ⏭️ No es factura válida (sin número de cliente)")
            return None
    
    # ========== NÚMERO DE CLIENTE ==========
    match = re.search(r'(\d{7}-\d)(?:\s*[\d\w])', texto_pagina)  # ⬅️ CAMBIO: [\d\w] en lugar de [A-Z]
    if match:
        datos["Numero_Cliente"] = match.group(1).strip()
    else:
        return None

    # ===== CUENTA PADRE =====
    match_padre = re.search(r'\n(\d{7})\s*\n', texto_pagina)
    if match_padre:
        datos["Cuenta_Padre"] = match_padre.group(1)

    # ========== NOMBRE CLIENTE ==========
    if datos["Numero_Cliente"]:
        # 🔥 NUEVO: Captura MÁS agresiva del nombre completo
        patron = rf'{re.escape(datos["Numero_Cliente"])}\s*\n?\s*([\d\w].*?)(?:\n(?:CL|KR|AC|DG|AV|AVENIDA|CALLE|CARRERA|AK|PZ|TV|VT|RA|AP|MNZ|ZN|BRR|SM|LT|MZ)\s+|\n\n|$)'
        match = re.search(patron, texto_pagina, re.DOTALL)
        if match:
            nombre = match.group(1).strip()
            # 🔥 NUEVO: Limpiar saltos de línea Y caracteres inválidos para nombres de archivo
            nombre = nombre.replace('\n', ' ')
            nombre = re.sub(r'\s+', ' ', nombre)
            nombre = re.sub(r'[\r\n\t]', '', nombre)  # ⬅️ AGREGAR ESTA LÍNEA
            if 3 < len(nombre) < 150:
                datos["Nombre_Cliente"] = nombre
    
    # ===== DIRECCIÓN =====
    if datos.get("Numero_Cliente"):
        direccion = extraer_direccion(texto_pagina, datos["Numero_Cliente"])
        if direccion:
            datos["Direccion"] = direccion
    
    # 🔥 NUEVO: corregir dirección con coordenadas si es necesario
    if pdf_path and numero_pagina and not datos.get("Direccion"):
        try:
            from pdfplumber import open as pdf_open
            with pdf_open(pdf_path) as pdf:
                page = pdf.pages[numero_pagina - 1]
                bbox = (100, 120, 500, 260)
                texto = page.within_bbox(bbox).extract_text()
                if texto and datos.get("Nombre_Cliente"):
                    lineas = [l.strip() for l in texto.split('\n') if l.strip()]
                    direccion_lineas = []
                    for l in lineas:
                        if datos.get("Nombre_Cliente") and datos["Nombre_Cliente"] in l:
                            continue
                        direccion_lineas.append(l)
                    if direccion_lineas:
                        direccion = " ".join(direccion_lineas)
                        datos["Direccion"] = re.sub(r'\s+', ' ', direccion).strip()
        except:
            pass
    
    # # ========== PERÍODO DE FACTURACIÓN ==========
    # match = re.search(
    #     r'(\d{1,2}\s+[A-Z]{3}/\d{4}\s+[AaBb]\s+\d{1,2}\s+[A-Z]{3}/\d{4})',
    #     texto_pagina
    # )
    # if match:
    #     datos["Periodo_Facturacion"] = match.group(1).strip()

    # ===== FECHAS =====
    fechas = re.findall(r'(\d{2}\s+[A-Z]{3}\s*/\d{4})', texto_pagina)
    if len(fechas) >= 2:
        datos["Fecha_Inicial"] = fechas[0]
        datos["Fecha_Final"] = fechas[1]
    
    # ========== TOTAL ENERGÍA ==========
    match_energia = re.search(r"(?i)(?:Total Energía y Otros|SUBTOTAL VALOR CONSUMO)\s*[\$]*\s*([\d\.,]+)", texto_pagina)
    if match_energia:
        datos["Total_Energia"] = limpiar_numero(match_energia.group(1))
    else:
        match_energia_alt = re.search(r"(?i)TOTAL\s*CONCEPTO\s*ENERG[IÍ]A\s*[\$]*\s*([\d\.,]+)", texto_pagina)
        datos["Total_Energia"] = limpiar_numero(match_energia_alt.group(1)) if match_energia_alt else 0.0

    # ===== TOTAL ASEO =====
    datos["Total_Aseo"] = extraer_total_aseo(texto_pagina)

    # ========== CONSUMO (opcional) ==========
    match_consumo = re.search(r'CONSUMO ACTIVA SENCILLA\s+(\d{1,3}(?:\.\d{3})*)', texto_pagina, re.IGNORECASE)
    if match_consumo:
        datos["Lectura_Actual"] = limpiar_numero(match_consumo.group(1))
    
    # =========================================
    # 🔥 ENERGÍA CONSUMIDA (MÉTODO DEFINITIVO)
    # =========================================
    for linea in texto_pagina.split('\n'):
        if 'CONSUMO ACTIVA SENCILLA' in linea.upper():

            numeros = re.findall(r'\d{1,3}(?:\.\d{3})*', linea)

            if len(numeros) >= 6:
                energia = numeros[4]  # 🔥 POSICIÓN CORRECTA

                datos["Energia_Consumida"] = limpiar_numero(energia)

                logging.info(f"   ⚡ Energía Consumida: {datos['Energia_Consumida']} kWh")

                break
    if not datos.get("Energia_Consumida"):
        logging.warning("   ⚠️ No se pudo extraer Energía Consumida")

    # ========== NIT EMPRESA ==========
    match = re.search(r'NIT\.\s+(\d{3}\.\d{3}\.\d{3}-\d)', texto_pagina)
    if match:
        datos["NIT_Empresa"] = match.group(1).strip()
    else:
        match = re.search(r'NIT[.:\s]+(\d{9,10}-\d)', texto_pagina)
        if match:
            datos["NIT_Empresa"] = match.group(1).strip()

    # ========== NIT ASEO ==========
    datos["NIT_Aseo"] = extraer_nit_aseo(texto_pagina)

    #Cuenta Contrato ASEO
    datos["Cuenta_Contrato"] = extraer_cuenta_contrato(
        texto_pagina,
        pdf_path=pdf_path,
        numero_pagina=numero_pagina
    )

    # 🔥 PERIODO ASEO (nuevo)
    periodo_aseo = extraer_periodo_aseo(texto_pagina)

    if periodo_aseo:
        datos["Periodo_Aseo"] = periodo_aseo

        # =========================================
        # 🔥 SEPARAR FECHAS ASEO (NUEVO)
        # =========================================
        match = re.search(r'(\d{2}/\d{2}/\d{4})\s*(?:al|-|a)\s*(\d{2}/\d{2}/\d{4})', periodo_aseo)

        if match:
            datos["Fecha_Inicial_Aseo"] = match.group(1)
            datos["Fecha_Final_Aseo"] = match.group(2)

            logging.info(f"   🧹 ASEO Desde: {datos['Fecha_Inicial_Aseo']} → {datos['Fecha_Final_Aseo']}")

    # ========== PRESTADOR ASEO ==========
    datos["Prestador_Aseo"] = extraer_prestador_aseo(texto_pagina)
    
    # 🔥 NUEVO: corregir prestador con coordenadas si es necesario
    if pdf_path and numero_pagina and not datos.get("Prestador_Aseo"):
        prestador = extraer_prestador_por_coordenadas(pdf_path, numero_pagina)
        if prestador:
            datos["Prestador_Aseo"] = prestador

    # ========== CONCEPTOS ASEO (CADA UNO EN SU COLUMNA) ==========
    conceptos = extraer_conceptos_aseo(texto_pagina)
    datos.update(conceptos)
    
    # ========== ESTADO DE PAGO ==========
    total_pagar = 0
    match_total_hoja = re.search(r"(?i)Total a Pagar\s*[\$]*\s*([\d\.,]+)", texto_pagina)
    if match_total_hoja:
        total_pagar = limpiar_numero(match_total_hoja.group(1)) or 0
    else:
        valores = re.findall(r'\$\s*(\d{1,3}(?:\.\d{3})+)', texto_pagina)
        valores_limpios = [limpiar_numero(v) for v in valores if limpiar_numero(v)]
        total_pagar = max(valores_limpios) if valores_limpios else 0.0
    
    datos["Total_Energia"] = total_pagar if not datos.get("Total_Energia") else datos["Total_Energia"]
    
    if total_pagar == 0 or total_pagar is None or "$0" in texto_pagina or "SALDO CRÉDITO" in texto_pagina:
        datos["Estado_Pago"] = "Sin Deuda"
    elif "PAGADO" in texto_pagina.upper():
        datos["Estado_Pago"] = "Pagado"
    else:
        datos["Estado_Pago"] = "Pendiente"
    
    # ===== COSTO UNITARIO (COORDENADAS - DEFINITIVO) =====
    if pdf_path and numero_pagina:
        costo = extraer_costo_unitario_por_coordenadas(pdf_path, numero_pagina)
        if costo:
            datos["Costo_Unitario"] = limpiar_numero(costo)
            logging.info(f"   → Costo Unitario (coord): ${costo}")
    
    return datos


# ==================== NUEVA FUNCIÓN: GENERAR EXCEL ====================
def generar_excel_facturas(datos_lista, output_path="resultado_facturas.xlsx"):
    """
    Genera un archivo Excel con todos los datos de facturas extraídas.
    Detecta columnas dinámicamente según los datos encontrados.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # 🔥 DETECTAR TODAS LAS COLUMNAS DINÁMICAMENTE
        todas_las_columnas = set(HEADERS_BASE)
        
        for datos in datos_lista:
            for clave in datos.keys():
                if clave not in todas_las_columnas:
                    todas_las_columnas.add(clave)
        
        # Ordenar: primero las base, luego las dinámicas ASEO
        headers_finales = list(HEADERS_BASE)
        aseo_dinamicos = sorted([col for col in todas_las_columnas if col.startswith('ASEO_') and col not in HEADERS_BASE])
        headers_finales.extend(aseo_dinamicos)
        
        logging.info(f"📋 Columnas detectadas: {len(headers_finales)}")
        logging.info(f"   Base: {len(HEADERS_BASE)}, Dinámicas ASEO: {len(aseo_dinamicos)}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturas"
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers_finales, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            
            if header.startswith('ASEO_') and header not in HEADERS_BASE:
                cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")  # Naranja
            else:
                cell.fill = PatternFill(start_color="1F4E38", end_color="1F4E38", fill_type="solid")  # Verde
            
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        
        for row_idx, datos in enumerate(datos_lista, 2):
            for col_idx, header in enumerate(headers_finales, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                valor = datos.get(header, "")
                cell.value = valor if valor is not None else ""
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = thin_border
        
        for col_idx in range(1, len(headers_finales) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16
        
        ws.freeze_panes = "A2"
        
        wb.save(output_path)
        logging.info(f"✅ Excel generado exitosamente: {output_path}")
        logging.info(f"   📊 Total filas: {len(datos_lista)}")
        logging.info(f"   📋 Total columnas: {len(headers_finales)}")
        return True
        
    except ImportError:
        logging.error("❌ openpyxl no instalado. Ejecuta: pip install openpyxl")
        return False
    except Exception as e:
        logging.error(f"❌ Error generando Excel: {e}")
        return False

# 🔥 PARÁMETROS DE LOTE
FACTURAS_POR_LOTE = 15
PAUSA_ENTRE_LOTES = 30


def main():
    """Función principal"""
    pdf_files = glob.glob(os.path.join(PDF_DIR, "*.pdf"))

    if not pdf_files:
        logging.warning(f"⚠️ No hay PDFs en: {PDF_DIR}")
        return

    logging.info(f"📊 Procesando {len(pdf_files)} archivo(s) PDF...")

    total_facturas = 0
    total_paginas = 0
    facturas_exitosas = 0
    facturas_fallidas = 0
    todas_las_facturas = []  # 🔥 NUEVA: almacenar todos los datos
    facturas_en_lote = 0
    num_lote = 1
    
    for pdf_file in pdf_files:
        logging.info(f"\n{'='*70}")
        logging.info(f"📄 PROCESANDO: {os.path.basename(pdf_file)}")
        logging.info(f"{'='*70}")
        
        paginas = extraer_paginas_pdf(pdf_file)
        
        if not paginas:
            logging.warning(f"⚠️ No se pudo extraer páginas")
            continue
        
        # Procesar CADA PÁGINA
        for pagina_info in paginas:
            num_pagina = pagina_info['numero']
            texto_pagina = pagina_info['texto']
            total_paginas += 1
            
            logging.info(f"\n📖 Página {num_pagina}/{len(paginas)}")
            
            # 🔥 AHORA PASAMOS pdf_path y numero_pagina
            datos = extraer_datos_factura(texto_pagina, pdf_path=pdf_file, numero_pagina=num_pagina)
            
            if datos and datos.get("Numero_Cliente"):
                total_facturas += 1
                facturas_exitosas += 1
                todas_las_facturas.append(datos)
                facturas_en_lote += 1
                
                logging.info(f"   ✅ FACTURA PROCESADA ({facturas_en_lote}/{FACTURAS_POR_LOTE}):")
                logging.info(f"      🔹 Cliente: {datos.get('Numero_Cliente')} - {datos.get('Nombre_Cliente')}")
                logging.info(f"      🔹 Período: {datos.get('Periodo_Facturacion')}")
                logging.info(f"      🔹 Total Energía: ${datos.get('Total_Energia')}")
                logging.info(f"      🔹 Total Aseo: ${datos.get('Total_Aseo')}")
                logging.info(f"      🔹 Consumo: {datos.get('Lectura_Actual')} {datos.get('Consumo_Unidad')}")
                logging.info(f"      🔹 NIT: {datos.get('NIT_Empresa')}")
                logging.info(f"      🔹 Prestador: {datos.get('Prestador_Aseo')}")
                
                # 🔥 PAUSA DESPUÉS DE CADA LOTE
                if facturas_en_lote >= FACTURAS_POR_LOTE:
                    logging.info(f"\n{'='*70}")
                    logging.info(f"⏸️  PAUSA ENTRE LOTES")
                    logging.info(f"📊 LOTE {num_lote}: {facturas_en_lote} facturas procesadas")
                    logging.info(f"⏳ Esperando {PAUSA_ENTRE_LOTES} segundos...")
                    logging.info(f"{'='*70}")

                    time.sleep(PAUSA_ENTRE_LOTES)

                    facturas_en_lote = 0
                    num_lote += 1

                    logging.info(f"✅ Reanudando procesamiento - LOTE {num_lote}\n")
            
            else:
                if datos is None:
                    logging.info(f"   ⏭️ Página {num_pagina}: No es una factura ENEL válida")
                else:
                    logging.info(f"   ⏭️ Página {num_pagina}: Sin número de cliente")
                facturas_fallidas += 1
        
        logging.info(f"\n✅ Archivo completado: {len(paginas)} páginas procesadas")
    
    # 🔥 NUEVA: Generar Excel al final
    logging.info(f"\n{'='*70}")
    logging.info(f"📊 GENERANDO ARCHIVO EXCEL...")
    logging.info(f"{'='*70}")
    
    if todas_las_facturas:
        output_file = os.path.join(PDF_DIR, "resultado_facturas.xlsx")
        if generar_excel_facturas(todas_las_facturas, output_file):
            logging.info(f"✅ Archivo guardado en: {output_file}")
        else:
            logging.error(f"❌ Error al generar Excel")
    else:
        logging.warning(f"⚠️ No hay facturas para exportar")
    
    # RESUMEN FINAL
    logging.info(f"\n{'='*70}")
    logging.info(f"✅ PROCESO FINALIZADO")
    logging.info(f"{'='*70}")
    logging.info(f"📊 RESUMEN GENERAL:")
    logging.info(f"   📄 Total de páginas procesadas: {total_paginas}")
    logging.info(f"   ✅ Total facturas extraídas: {total_facturas}")
    logging.info(f"   ❌ Facturas con error: {facturas_fallidas}")
    logging.info(f"📁 Archivo Excel: resultado_facturas.xlsx")
    logging.info(f"📝 Log disponible en: {LOG_FILE}")
    logging.info(f"{'='*70}\n")


if __name__ == "__main__":
    main()