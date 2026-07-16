# -*- coding: utf-8 -*-
"""
Contrafirma masiva DocuSeal.

Logica: el jefe firma UNA sola vez; su firma se estampa (rol completado) en N copias
del documento, y cada copia se envia a un empleado para que firme la suya.
El PDF final y el certificado de cada copia incluyen a AMBAS partes.

Origen de la firma del jefe (elige uno):
  --firma-de-submission 123   toma la firma de un envio que el jefe YA firmo en DocuSeal
  --firma-imagen firma.png    usa una imagen PNG/JPG de la firma

Uso tipico:
  1) Enviar el documento maestro al jefe desde la web y esperar a que firme.
  2) python contrafirma_masiva.py --archivo empleados.csv --template-id 5 ^
       --jefe-email jefe@copservir.com --firma-de-submission 123

La plantilla debe tener DOS roles: el primero = jefe, el segundo = empleado
(o indica cuales con --rol-jefe / --rol-empleado).
"""
import argparse
import base64
import csv
import json
import mimetypes
import os
import re
import sys
import time
import urllib.request
import urllib.error

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-']+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
COL_EMAIL = ('email', 'correo', 'correo electronico', 'correo electrónico', 'e-mail', 'mail')
COL_NOMBRE = ('nombre', 'name', 'nombres', 'empleado')

OPENER = urllib.request.build_opener(urllib.request.ProxyHandler({}))  # sin proxy corporativo


def api(url_base, token, metodo, ruta, cuerpo=None):
    req = urllib.request.Request(
        url_base.rstrip('/') + ruta,
        method=metodo,
        data=json.dumps(cuerpo).encode('utf-8') if cuerpo is not None else None,
        headers={'X-Auth-Token': token, 'Content-Type': 'application/json'},
    )
    with OPENER.open(req, timeout=60) as resp:
        return json.loads(resp.read().decode('utf-8'))


def descargar_base64(url):
    with OPENER.open(url, timeout=60) as resp:
        tipo = resp.headers.get('Content-Type', 'image/png').split(';')[0]
        return f'data:{tipo};base64,' + base64.b64encode(resp.read()).decode()


def leer_lista(ruta):
    ext = os.path.splitext(ruta)[1].lower()
    if ext == '.xlsx':
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit('Para .xlsx instala openpyxl o exporta como CSV.')
        hoja = load_workbook(ruta, read_only=True).active
        filas = [[('' if c is None else str(c)) for c in fila] for fila in hoja.iter_rows(values_only=True)]
    else:
        with open(ruta, encoding='utf-8-sig', newline='') as f:
            try:
                dialecto = csv.Sniffer().sniff(f.read(4096), delimiters=',;\t')
            except csv.Error:
                dialecto = csv.excel
            f.seek(0)
            filas = [list(fila) for fila in csv.reader(f, dialecto)]
    if not filas:
        sys.exit(f'El archivo {ruta} esta vacio.')

    encabezado = [c.strip().lower() for c in filas[0]]
    idx_email = next((i for i, c in enumerate(encabezado) if c in COL_EMAIL), None)
    idx_nombre = next((i for i, c in enumerate(encabezado) if c in COL_NOMBRE), None)
    vistos, lista = set(), []

    def agregar(email, nombre=None):
        email = email.strip().lower()
        if email and email not in vistos and EMAIL_RE.fullmatch(email):
            vistos.add(email)
            lista.append((email, (nombre or '').strip() or None))

    if idx_email is not None:
        for fila in filas[1:]:
            if len(fila) > idx_email:
                nombre = fila[idx_nombre] if idx_nombre is not None and len(fila) > idx_nombre else None
                agregar(fila[idx_email], nombre)
    else:
        for fila in filas:
            for celda in fila:
                for email in EMAIL_RE.findall(celda):
                    agregar(email)
    return lista


def obtener_firma(args):
    """Devuelve la firma del jefe como data:image base64."""
    if args.firma_imagen:
        tipo = mimetypes.guess_type(args.firma_imagen)[0] or 'image/png'
        with open(args.firma_imagen, 'rb') as f:
            return f'data:{tipo};base64,' + base64.b64encode(f.read()).decode()

    sub = api(args.url, args.token, 'GET', f'/api/submissions/{args.firma_de_submission}')
    for firmante in sub['submitters']:
        if args.jefe_email and firmante['email'].lower() != args.jefe_email.lower():
            continue
        if not firmante.get('completed_at'):
            continue
        for v in firmante.get('values', []):
            valor = v.get('value')
            if isinstance(valor, str) and valor.startswith('http'):
                print(f"Firma tomada del envio {args.firma_de_submission} (firmante {firmante['email']}).")
                return descargar_base64(valor)
    sys.exit(f'No se encontro una firma completada del jefe en la submission {args.firma_de_submission}.')


def main():
    p = argparse.ArgumentParser(description='Contrafirma masiva: el jefe firma una vez, N empleados firman su copia.')
    p.add_argument('--archivo', required=True, help='CSV/TXT/XLSX con los correos de empleados')
    p.add_argument('--template-id', required=True, type=int, help='Plantilla con DOS roles (jefe y empleado)')
    p.add_argument('--jefe-email', required=True, help='Correo del jefe (queda en el certificado)')
    p.add_argument('--jefe-nombre', default=None, help='Nombre del jefe')
    p.add_argument('--firma-de-submission', type=int, default=None, help='ID del envio donde el jefe ya firmo')
    p.add_argument('--firma-imagen', default=None, help='Ruta a imagen de la firma (PNG/JPG)')
    p.add_argument('--rol-jefe', default=None, help='Nombre del rol del jefe (default: primer rol)')
    p.add_argument('--rol-empleado', default=None, help='Nombre del rol del empleado (default: segundo rol)')
    p.add_argument('--url', default='http://localhost:3000')
    p.add_argument('--token', default=os.environ.get('DOCUSEAL_TOKEN'))
    p.add_argument('--sin-email', action='store_true', help='Crear envios SIN mandar correos (prueba)')
    p.add_argument('--mensaje', default=None, help='Mensaje del correo de invitacion')
    p.add_argument('--si', action='store_true', help='No pedir confirmacion')
    args = p.parse_args()

    if not args.token:
        sys.exit('Falta el token: --token o variable DOCUSEAL_TOKEN.')
    if not args.firma_de_submission and not args.firma_imagen:
        sys.exit('Indica el origen de la firma del jefe: --firma-de-submission ID o --firma-imagen ruta.')

    lista = leer_lista(args.archivo)
    if not lista:
        sys.exit('No se encontro ningun correo valido en el archivo.')

    plantilla = api(args.url, args.token, 'GET', f'/api/templates/{args.template_id}')
    roles = [s['name'] for s in plantilla['submitters']]
    rol_jefe = args.rol_jefe or roles[0]
    rol_empleado = args.rol_empleado or next((r for r in roles if r != rol_jefe), None)
    if rol_jefe not in roles or not rol_empleado or rol_empleado not in roles:
        sys.exit(f'Roles de la plantilla: {roles}. Revisa --rol-jefe / --rol-empleado.')

    # uuid del rol jefe -> prellenar sus campos de firma/imagen con la firma capturada
    uuid_jefe = next(s['uuid'] for s in plantilla['submitters'] if s['name'] == rol_jefe)
    campos_firma_jefe = [f['uuid'] for f in plantilla['fields']
                         if f.get('submitter_uuid') == uuid_jefe
                         and f.get('type') in ('signature', 'initials', 'image', 'stamp')]
    if not campos_firma_jefe:
        sys.exit(f"El rol '{rol_jefe}' no tiene campos de firma en la plantilla.")

    firma = obtener_firma(args)

    print(f"Plantilla {args.template_id}: \"{plantilla['name']}\"")
    print(f"  Jefe:     rol '{rol_jefe}' <{args.jefe_email}> -> llega FIRMADO ({len(campos_firma_jefe)} campo(s))")
    print(f"  Empleado: rol '{rol_empleado}' -> {len(lista)} copias por firmar")
    print(f"  Enviar email: {'NO (prueba)' if args.sin_email else 'SI'}")
    if not args.si and input('Continuar? (s/N): ').strip().lower() != 's':
        sys.exit('Cancelado.')

    jefe = {'role': rol_jefe, 'email': args.jefe_email, 'completed': True,
            'values': {uuid: firma for uuid in campos_firma_jefe}}
    if args.jefe_nombre:
        jefe['name'] = args.jefe_nombre

    ok, errores = 0, []
    for n, (email, nombre) in enumerate(lista, 1):
        empleado = {'role': rol_empleado, 'email': email}
        if nombre:
            empleado['name'] = nombre
        cuerpo = {
            'template_id': args.template_id,
            'send_email': not args.sin_email,
            'submitters': [jefe, empleado],
        }
        if args.mensaje:
            cuerpo['message'] = {'body': args.mensaje}
        try:
            api(args.url, args.token, 'POST', '/api/submissions', cuerpo)
            ok += 1
        except urllib.error.HTTPError as e:
            errores.append((email, f'HTTP {e.code}: {e.read().decode("utf-8", "replace")[:200]}'))
        except Exception as e:
            errores.append((email, str(e)))
        if n % 25 == 0 or n == len(lista):
            print(f'  {n}/{len(lista)} procesados ({ok} ok, {len(errores)} errores)')
        time.sleep(0.1)

    print(f'\nResultado: {ok} copias creadas con la firma del jefe estampada, {len(errores)} errores.')
    if errores:
        ruta_err = os.path.join(os.path.dirname(os.path.abspath(args.archivo)), 'errores_contrafirma.csv')
        with open(ruta_err, 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.writer(f)
            w.writerow(['email', 'error'])
            w.writerows(errores)
        print(f'Detalle de errores en: {ruta_err}')


if __name__ == '__main__':
    main()
