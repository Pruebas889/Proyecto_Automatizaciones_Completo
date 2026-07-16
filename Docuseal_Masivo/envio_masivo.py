# -*- coding: utf-8 -*-
"""
Envio masivo de documentos DocuSeal.

Regla: un formato (plantilla), N empleados -> N copias independientes,
cada empleado firma la suya.

Uso:
  python envio_masivo.py --archivo lista.csv --template-id 1
  python envio_masivo.py --archivo lista.csv --template-id 1 --sin-email   (prueba: no manda correos)
  python envio_masivo.py --archivo lista.xlsx --template-id 5 --url https://firmasdigitales.copservir.com --token XXXX

Archivo de lista aceptado:
  .csv / .txt : extrae todo lo que tenga formato de correo (sirve pegar la columna de Excel en un .txt)
                Si el CSV tiene encabezados 'email/correo' y 'nombre/name', usa esas columnas.
  .xlsx       : requiere openpyxl (pip install openpyxl); usa las mismas columnas.

Token: se pasa con --token o con la variable de entorno DOCUSEAL_TOKEN.
       Se genera en DocuSeal: Configuracion -> API.
"""
import argparse
import csv
import json
import os
import re
import sys
import time
import urllib.request
import urllib.error

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-']+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")

COL_EMAIL = ('email', 'correo', 'correo electronico', 'correo electrónico', 'e-mail', 'mail')
COL_NOMBRE = ('nombre', 'name', 'nombres', 'empleado')


def api(url_base, token, metodo, ruta, cuerpo=None):
    req = urllib.request.Request(
        url_base.rstrip('/') + ruta,
        method=metodo,
        data=json.dumps(cuerpo).encode('utf-8') if cuerpo is not None else None,
        headers={'X-Auth-Token': token, 'Content-Type': 'application/json'},
    )
    # sin proxy: el FortiProxy corporativo no debe interceptar llamadas a localhost/red interna
    opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
    with opener.open(req, timeout=60) as resp:
        return json.loads(resp.read().decode('utf-8'))


def leer_lista(ruta):
    """Devuelve lista de (email, nombre_o_None) sin duplicados, en orden."""
    ext = os.path.splitext(ruta)[1].lower()
    filas = []
    if ext == '.xlsx':
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("Para .xlsx instala openpyxl (pip install openpyxl) o exporta la hoja como CSV.")
        hoja = load_workbook(ruta, read_only=True).active
        filas = [[('' if c is None else str(c)) for c in fila] for fila in hoja.iter_rows(values_only=True)]
    else:
        with open(ruta, encoding='utf-8-sig', newline='') as f:
            try:
                dialecto = csv.Sniffer().sniff(f.read(4096), delimiters=',;\t')
            except csv.Error:
                dialecto = csv.excel
            f.seek(0)
            filas = [list(fila) for fila in csv.reader(f, dialecto)]

    if not filas:
        sys.exit(f"El archivo {ruta} esta vacio.")

    # ubicar columnas por encabezado; si no hay, extraer correos por regex de todo el contenido
    encabezado = [c.strip().lower() for c in filas[0]]
    idx_email = next((i for i, c in enumerate(encabezado) if c in COL_EMAIL), None)
    idx_nombre = next((i for i, c in enumerate(encabezado) if c in COL_NOMBRE), None)

    vistos, lista = set(), []

    def agregar(email, nombre=None):
        email = email.strip().lower()
        if email and email not in vistos and EMAIL_RE.fullmatch(email):
            vistos.add(email)
            lista.append((email, (nombre or '').strip() or None))

    if idx_email is not None:
        for fila in filas[1:]:
            if len(fila) > idx_email:
                nombre = fila[idx_nombre] if idx_nombre is not None and len(fila) > idx_nombre else None
                agregar(fila[idx_email], nombre)
    else:
        for fila in filas:
            for celda in fila:
                for email in EMAIL_RE.findall(celda):
                    agregar(email)
    return lista


def main():
    p = argparse.ArgumentParser(description='Envio masivo DocuSeal: una copia por empleado.')
    p.add_argument('--archivo', required=True, help='CSV/TXT/XLSX con los correos')
    p.add_argument('--template-id', required=True, type=int, help='ID de la plantilla en DocuSeal')
    p.add_argument('--url', default='http://localhost:3000', help='URL de DocuSeal (default: local)')
    p.add_argument('--token', default=os.environ.get('DOCUSEAL_TOKEN'), help='Token de API (o var DOCUSEAL_TOKEN)')
    p.add_argument('--sin-email', action='store_true', help='Crear los envios SIN mandar correos (prueba)')
    p.add_argument('--mensaje', default=None, help='Mensaje opcional del correo de invitacion')
    p.add_argument('--si', action='store_true', help='No pedir confirmacion (para automatizaciones)')
    p.add_argument('--rol', default=None, help='Rol al que se asigna cada correo (default: primer rol de la plantilla)')
    args = p.parse_args()

    if not args.token:
        sys.exit('Falta el token: usa --token o define DOCUSEAL_TOKEN.')

    lista = leer_lista(args.archivo)
    if not lista:
        sys.exit('No se encontro ningun correo valido en el archivo.')

    plantilla = api(args.url, args.token, 'GET', f'/api/templates/{args.template_id}')
    roles = [s['name'] for s in plantilla['submitters']]
    print(f"Plantilla {args.template_id}: \"{plantilla['name']}\" | roles: {', '.join(roles)}")
    if args.rol:
        if args.rol not in roles:
            sys.exit(f"El rol '{args.rol}' no existe en la plantilla. Roles: {roles}")
        roles = [args.rol] + [r for r in roles if r != args.rol]
    if len(roles) > 1:
        print(f"ADVERTENCIA: la plantilla tiene {len(roles)} roles. Cada copia se creara solo con el rol")
        print(f"'{roles[0]}' asignado; los demas roles quedarian sin firmante. Para la regla de")
        print('"cada empleado firma su copia" la plantilla debe tener UN solo rol.')
        if not args.si and input('Continuar de todas formas? (s/N): ').strip().lower() != 's':
            sys.exit('Cancelado.')

    print(f"Correos a enviar: {len(lista)} | enviar email: {'NO (prueba)' if args.sin_email else 'SI'}")

    ok, errores = 0, []
    for n, (email, nombre) in enumerate(lista, 1):
        firmante = {'role': roles[0], 'email': email}
        if nombre:
            firmante['name'] = nombre
        cuerpo = {
            'template_id': args.template_id,
            'send_email': not args.sin_email,
            'submitters': [firmante],
        }
        if args.mensaje:
            cuerpo['message'] = {'body': args.mensaje}
        try:
            api(args.url, args.token, 'POST', '/api/submissions', cuerpo)
            ok += 1
        except urllib.error.HTTPError as e:
            errores.append((email, f'HTTP {e.code}: {e.read().decode("utf-8", "replace")[:200]}'))
        except Exception as e:  # red caida, timeout, etc.
            errores.append((email, str(e)))
        if n % 25 == 0 or n == len(lista):
            print(f'  {n}/{len(lista)} procesados ({ok} ok, {len(errores)} errores)')
        time.sleep(0.1)  # no saturar el servidor

    print(f'\nResultado: {ok} envios creados, {len(errores)} errores.')
    if errores:
        ruta_err = os.path.join(os.path.dirname(os.path.abspath(args.archivo)), 'errores_envio.csv')
        with open(ruta_err, 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.writer(f)
            w.writerow(['email', 'error'])
            w.writerows(errores)
        print(f'Detalle de errores en: {ruta_err}')


if __name__ == '__main__':
    main()
