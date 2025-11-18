# coding: utf-8
import os
import re
import sys
import glob
import io
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import PyPDF2
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseUpload, MediaFileUpload

# Configuración de logs
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

# Forzar codificación UTF-8 para stdout
if sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)

# Configuración de Google Drive
SCOPES = ['https://www.googleapis.com/auth/drive.file']
FOLDER_ID_GOOGLE_DRIVE = '14U16Cej8SroNhWfAfL102gNHs-qtS_GY'
CREDENTIALS_FILE = 'credentials.json'
PDF_DIR = r"C:\Users\dforero\Downloads\Facturas_Claro_Pruebas"  # Definición de PDF_DIR

def get_or_create_pdf_folder(service, parent_folder_id):
    """Busca o crea una subcarpeta llamada 'PDFs' dentro de la carpeta principal."""
    try:
        query = f"'{parent_folder_id}' in parents and name = 'PDFs' and mimeType = 'application/vnd.google-apps.folder' and trashed = false"
        results = service.files().list(
            q=query,
            spaces='drive',
            fields='files(id, name)',
            pageSize=1
        ).execute()
        folders = results.get('files', [])
        if folders:
            pdf_folder_id = folders[0]['id']
            logging.info(f"Subcarpeta 'PDFs' encontrada con ID: {pdf_folder_id} 📁")
            return pdf_folder_id
        folder_metadata = {
            'name': 'PDFs',
            'mimeType': 'application/vnd.google-apps.folder',
            'parents': [parent_folder_id]
        }
        folder = service.files().create(body=folder_metadata, fields='id').execute()
        pdf_folder_id = folder.get('id')
        logging.info(f"Subcarpeta 'PDFs' creada con ID: {pdf_folder_id} 📁")
        return pdf_folder_id
    except HttpError as error:
        logging.error(f"Error al buscar o crear la subcarpeta 'PDFs': {error} ❌")
        return None
    except Exception as e:
        logging.error(f"Error inesperado al buscar o crear la subcarpeta 'PDFs': {e} ❌")
        return None

def authenticate_google_drive():
    creds = None
    token_path = 'token.json'
    
    # Intentar cargar el token existente
    if os.path.exists(token_path):
        try:
            creds = Credentials.from_authorized_user_file(token_path, SCOPES)
            # Verificar si el token es válido refrescándolo
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
        except Exception as e:
            logging.warning(f"Token inválido o revocado: {e}. Eliminando token.json para reautenticar.")
            os.remove(token_path)
            creds = None

    # Si no hay credenciales válidas, autenticar de nuevo
    if not creds or not creds.valid:
        try:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
            with open(token_path, 'w') as token:
                token.write(creds.to_json())
            logging.info("Nuevo token generado y guardado en token.json. ✅")
        except Exception as e:
            logging.error(f"Error al autenticar con Google Drive: {e} ❌")
            raise
    return creds

def upload_file_to_drive(file_obj, filename, folder_id, is_file_path=False):
    try:
        creds = authenticate_google_drive()
        service = build('drive', 'v3', credentials=creds)
        target_folder_id = folder_id
        if is_file_path:
            pdf_folder_id = get_or_create_pdf_folder(service, folder_id)
            if not pdf_folder_id:
                logging.error(f"No se pudo obtener o crear la subcarpeta 'PDFs' para '{filename}'")
                return None
            target_folder_id = pdf_folder_id
        file_metadata = {'name': filename, 'parents': [target_folder_id]}
        if is_file_path:
            media = MediaFileUpload(file_obj, mimetype='application/pdf', resumable=True)
        else:
            media = MediaIoBaseUpload(file_obj, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
        file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        logging.info(f"Archivo '{filename}' subido a Google Drive en carpeta ID: {target_folder_id}. ID: {file.get('id')} ✅")
        return file.get('id')
    except HttpError as error:
        logging.error(f"Error al subir '{filename}' a Google Drive: {error} ❌")
        return None
    except Exception as e:
        logging.error(f"Error inesperado al subir '{filename}' a Google Drive: {e} ❌")
        return None

def extraer_texto_pdf(pdf_path):
    try:
        with open(pdf_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            texto = ""
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    texto += page_text
            return texto
    except Exception as e:
        logging.error(f"Error al leer el PDF '{os.path.basename(pdf_path)}': {e} ❌")
        return None

def extraer_datos(texto):
    datos = {
        "Referencia de Pago": None,
        "Nº de Identificación": None,
        "Factura Electrónica de Venta": None,
        "Fecha de Factura": None,
        "Periodo de Facturación": None,
        "Pagar Antes De": None,
        "Total a Pagar": None,
        "Pagos Efectuados": None,
        "Internet Total": None,
        "Telefonia Total": None,
        "IVA": 0.0,
        "Servicios Detalle": [],
        "Subtotal Servicios (Base Imponible)": 0.0,
        "IVA Calculado": 0.0,
        "Diferencia IVA": None,
        "IVA Correcto": "No aplica"
    }

    patrones = {
        "Referencia de Pago": r"REFERENCIA DE PAGO:\s*([A-Za-z0-9]+)",
        "Nº de Identificación": r"N[°ºoO] DE IDENTIFICACIÓN:\s*([A-Za-z0-9\-\. ]+)",
        "Factura Electrónica de Venta": r"FACTURA ELECTRÓNICA DE VENTA:\s*([A-Za-z0-9 ]+)",
        "Fecha de Factura": r"FECHA DE FACTURA:\s*([A-Za-z]{3} \d{2}/\d{2})",
        "Periodo de Facturación": r"PERIODO DE FACTURACIÓN:\s*([A-Za-z]{3} \d{2}/\d{2} - [A-Za-z]{3} \d{2}/\d{2})",
        "Pagar Antes De": r"PAGAR ANTES DE:\s*([A-Za-z0-9 /]+)",
        "Total a Pagar": r"TOTAL A PAGAR:\s*\$?\s*([\d\.,]+)",
        "Pagos Efectuados": r"Pagos efectuados\s*\$?\s*(-?[\d\.,]+)"
    }

    for key, pat in patrones.items():
        match = re.search(pat, texto, re.IGNORECASE)
        if match:
            value_str = match.group(1)
            if key in ["Total a Pagar", "Pagos Efectuados"]:
                try:
                    datos[key] = float(value_str.replace('.', '').replace(',', '.'))
                except ValueError:
                    datos[key] = None
            else:
                datos[key] = value_str.strip()

    # Extraer total de Internet: buscar el valor asociado a "INTERNET" como total (sin fechas)
    match_internet_total = re.search(r"INTERNET\s+(?!.*\d{2}-[A-Za-z]{3}-\d{2,4})\$?\s*([\d\.,]+)", texto, re.IGNORECASE)
    if match_internet_total:
        try:
            datos["Internet Total"] = float(match_internet_total.group(1).replace('.', '').replace(',', '.'))
        except ValueError:
            pass

    match_telefonia_total = re.search(r"TELEFONÍA\s+\$?\s*([\d\.,]+)", texto, re.IGNORECASE)
    if match_telefonia_total:
        try:
            datos["Telefonia Total"] = float(match_telefonia_total.group(1).replace('.', '').replace(',', '.'))
        except ValueError:
            pass

    iva_matches = re.findall(r"\bIVA\b\s*\$?\s*([\d\.,]+)", texto, re.IGNORECASE)
    total_iva_extraido = 0.0
    for m in iva_matches:
        try:
            total_iva_extraido += float(m.replace('.', '').replace(',', '.'))
        except ValueError:
            pass
    datos["IVA"] = round(total_iva_extraido, 2)

    patron_servicio_fecha = r"([A-ZÁÉÍÓÚÑ0-9 \*\+\-]{5,})\s+(\d{2}-[A-Za-z]{3}-\d{2,4})\s+(\d{2}-[A-Zaelz]{3}-\d{2,4})\s+(\d+)\s+\$?\s*([\-\d\.,]+)"
    for m in re.finditer(patron_servicio_fecha, texto):
        try:
            descripcion = m.group(1).strip()
            fecha_inicial = m.group(2)
            fecha_final = m.group(3)
            dias = m.group(4)
            valor = float(m.group(5).replace('.', '').replace(',', '.'))
            datos["Servicios Detalle"].append({
                "Descripción": descripcion,
                "Fecha Inicial": fecha_inicial,
                "Fecha Final": fecha_final,
                "Días": dias,
                "Valor": valor
            })
        except:
            continue

    patron_otro_servicio_simple = r"^(?!Descripción\s+Valor$)(?!.+IVA$)(?!.+DEUDA ANTERIOR)(.+?)\s+\$?\s*([\-\d\.,]+)$"
    otros_servicios_section_match = re.search(r"OTROS SERVICIOS\n\n(.*?)(?=\n\n|\Z)", texto, re.DOTALL)
    if otros_servicios_section_match:
        section_text = otros_servicios_section_match.group(1)
        for m in re.finditer(patron_otro_servicio_simple, section_text, re.MULTILINE):
            desc = m.group(1).strip()
            val_str = m.group(2)
            if "IVA" not in desc.upper() and "TOTAL A PAGAR" not in desc.upper() and "Importante:" not in desc and "REFERENCIA DE PAGO" not in desc:
                try:
                    valor = float(val_str.replace('.', '').replace(',', '.'))
                    if not any(s['Descripción'] == desc and s['Valor'] == valor for s in datos["Servicios Detalle"]):
                        datos["Servicios Detalle"].append({
                            "Descripción": desc,
                            "Fecha Inicial": "", "Fecha Final": "", "Días": "", "Valor": valor
                        })
                except ValueError:
                    continue

    subtotal_calculable_iva = 0.0
    for s in datos["Servicios Detalle"]:
        desc_upper = s["Descripción"].strip().upper()
        if not desc_upper.startswith("IVA") and not desc_upper.startswith("DEUDA ANTERIOR"):
            subtotal_calculable_iva += s["Valor"]
    datos["Subtotal Servicios (Base Imponible)"] = subtotal_calculable_iva

    iva_extraido = datos.get("IVA", 0.0)
    if subtotal_calculable_iva <= 0:
        datos["IVA Calculado"] = 0.0
        datos["Diferencia IVA"] = 0.0
        datos["IVA Correcto"] = "Sí (No aplica base imponible)"
    elif iva_extraido > 0:
        datos["IVA Calculado"] = iva_extraido
        datos["Diferencia IVA"] = 0.0
        datos["IVA Correcto"] = "Sí (Desglosado en factura)"
    elif iva_extraido == 0 and subtotal_calculable_iva > 0:
        iva_calculado = round(subtotal_calculable_iva * 0.19, 2)
        datos["IVA Calculado"] = iva_calculado
        datos["Diferencia IVA"] = iva_calculado
        datos["IVA Correcto"] = "⚠ Revisión requerida (IVA = 0 y base > 0)"
    else:
        datos["IVA Calculado"] = 0.0
        datos["Diferencia IVA"] = 0.0
        datos["IVA Correcto"] = "No"

    # --- NUEVO: Calcular Internet y Telefonía SIN IVA ---
    def quitar_iva(monto, tasa=0.19):
        try:
            if monto is None:
                return None
            if monto == 0:
                return 0.0
            return round(monto / (1 + tasa), 2)
        except:
            return None

    # Sumar desde Servicios Detalle si hay líneas con INTERNET / TELEFONIA
    internet_raw = 0.0
    telefonia_raw = 0.0
    internet_found = False
    telefonia_found = False
    for s in datos["Servicios Detalle"]:
        d = s.get("Descripción", "").upper()
        val = s.get("Valor", 0.0)
        if "INTERNET" in d:
            internet_raw += val
            internet_found = True
        elif "TELEFONIA" in d or "TELÉFON" in d or "TELEFONÍA" in d:
            telefonia_raw += val
            telefonia_found = True

    # Si no se encontraron líneas, usar los valores extraídos por regex (si existen)
    if not internet_found and datos.get("Internet Total") is not None:
        internet_raw = datos["Internet Total"]
    if not telefonia_found and datos.get("Telefonia Total") is not None:
        telefonia_raw = datos["Telefonia Total"]

    # Guardar bruto y luego reemplazar por valor sin IVA
    datos["Internet Total (Bruto)"] = internet_raw if internet_raw != 0.0 else (datos.get("Internet Total") or None)
    datos["Telefonia Total (Bruto)"] = telefonia_raw if telefonia_raw != 0.0 else (datos.get("Telefonia Total") or None)

    datos["Internet Total"] = quitar_iva(internet_raw) if internet_raw else None
    datos["Telefonia Total"] = quitar_iva(telefonia_raw) if telefonia_raw else None
    # --- FIN NUEVO ---

    return datos

def exportar_excel_multiple(lista_datos, nombre_archivo, total_general, folder_id_drive):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Recibos Claro" # type: ignore
    ws_errores = wb.create_sheet("IVA Incorrecto")

    encabezado_fijo = [
        "Referencia de Pago", "Nº de Identificación", "Factura Electrónica de Venta",
        "Fecha de Factura", "Periodo de Facturación", "Pagar Antes De",
        "Total a Pagar", "Pagos Efectuados", "Internet", "Telefonía",
        "IVA (Extraído Total)", "Subtotal Servicios (Base Imponible)", "IVA Calculado (19%)", "Diferencia IVA", "¿IVA Correcto?"
    ]

    max_otros_servicios = 0
    for d in lista_datos:
        otros_servicios = [
            s for s in d.get("Servicios Detalle", [])
            if "INTERNET" not in s.get("Descripción", "").upper() and "TELEFONIA" not in s.get("Descripción", "").upper()
        ]
        if len(otros_servicios) > max_otros_servicios:
            max_otros_servicios = len(otros_servicios)

    encabezado_otros_servicios = []
    for i in range(1, max_otros_servicios + 1):
        encabezado_otros_servicios += [f"Servicio Adicional {i} Descripción", f"Servicio Adicional {i} Valor"]

    encabezado = encabezado_fijo + encabezado_otros_servicios

    for hoja in [ws, ws_errores]:
        hoja.append(encabezado)
        for col in range(1, len(encabezado)+1):
            cell = hoja.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

    for datos in lista_datos:
        internet_valor = datos.get("Internet Total")
        telefonica_valor = datos.get("Telefonia Total")
        
        # Restar 1 a Internet antes de subir a Excel
        if internet_valor is not None:
            internet_valor = internet_valor - 1
        
        otros_servicios_fila = []
        current_category = None  # Para rastrear la sección actual (Internet o Telefonía)

        for s in datos.get("Servicios Detalle", []):
            desc = s.get("Descripción", "").strip().upper()
            valor = s.get("Valor")
            # Detectar y cambiar categoría basada en palabras clave
            if "INTERNET" in desc:
                current_category = "internet"
            elif "TELEFONICA" in desc or "TELEFONIA" in desc or "TELÉFON" in desc:
                current_category = "telefonica"

            # Sumar a la categoría actual si aplica
            if current_category == "internet":
                pass  # Ya tenemos el total del recibo, no sumamos aquí
            elif current_category == "telefonica":
                pass  # Ya tenemos el total del recibo, no sumamos aquí
            else:
                otros_servicios_fila.append(s)

        fila = [
            datos.get("Referencia de Pago"),
            datos.get("Nº de Identificación"),
            datos.get("Factura Electrónica de Venta"),
            datos.get("Fecha de Factura"),
            datos.get("Periodo de Facturación"),
            datos.get("Pagar Antes De"),
            datos.get("Total a Pagar"),
            datos.get("Pagos Efectuados"),
            internet_valor,
            telefonica_valor,
            datos.get("IVA"),
            datos.get("Subtotal Servicios (Base Imponible)"),
            datos.get("IVA Calculado"),
            datos.get("Diferencia IVA"),
            datos.get("IVA Correcto")
        ]

        # Agregar servicios adicionales (si hay)
        for i in range(max_otros_servicios):
            if i < len(otros_servicios_fila):
                fila += [
                    otros_servicios_fila[i].get("Descripción"),
                    otros_servicios_fila[i].get("Valor")
                ]
            else:
                fila += ["", ""]

        ws.append(fila) # type: ignore
        if datos["IVA Correcto"] not in ["Sí", "Sí (No aplica base imponible)", "Sí (Desglosado en factura)"]:
            ws_errores.append(fila)

    ws.append([]) # type: ignore
    total_col_index = encabezado.index("Total a Pagar")
    total_row = [""] * len(encabezado)
    total_row[total_col_index] = total_general
    total_row[0] = "TOTAL GENERAL"
    ws.append(total_row) # type: ignore

    for hoja in [ws, ws_errores]:
        for i, col in enumerate(hoja.columns, 1):
            max_length = 0
            column_values = [str(cell.value) for cell in col if cell.value is not None]
            if column_values:
                max_length = max(len(val) for val in column_values)
            hoja.column_dimensions[get_column_letter(i)].width = (max_length + 2) * 1.2

    # Columnas sin decimales
    columnas_sin_decimales = ["Total a Pagar", "Pagos Efectuados", "Internet", "Telefonía", "IVA (Extraído Total)"]
    # Columnas con decimales
    formato_numero = '#,##0.00'
    formato_sin_decimales = '#,##0'
    columnas_numericas = [
        "Total a Pagar", "Pagos Efectuados", "Internet", "Telefonía",
        "IVA (Extraído Total)", "Subtotal Servicios (Base Imponible)",
        "IVA Calculado (19%)", "Diferencia IVA"
    ]
    for i in range(1, max_otros_servicios + 1):
        columnas_numericas.append(f"Servicio Adicional {i} Valor")

    for hoja in [ws, ws_errores]:
        for col_idx, col_name in enumerate(encabezado, 1):
            if col_name in columnas_numericas:
                for row in range(2, hoja.max_row + 1):
                    cell = hoja.cell(row=row, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        if col_name in columnas_sin_decimales:
                            cell.number_format = formato_sin_decimales
                        else:
                            cell.number_format = formato_numero

    excel_buffer = io.BytesIO()
    try:
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        logging.info(f"Excel '{nombre_archivo}' generado en memoria. ✅")
        uploaded_file_id = upload_file_to_drive(excel_buffer, nombre_archivo, folder_id_drive)
        if uploaded_file_id:
            logging.info(f"Archivo '{nombre_archivo}' subido exitosamente a Google Drive. ID: {uploaded_file_id} ✅")
        else:
            logging.error(f"Falló la subida de '{nombre_archivo}' a Google Drive. ❌")
    except Exception as e:
        logging.error(f"Error al generar o subir Excel: {e} ❌")

def main():
    pdf_files = glob.glob(os.path.join(PDF_DIR, "*.pdf"))

    if not pdf_files:
        logging.warning(f"No se encontraron archivos PDF en la carpeta: {PDF_DIR} 🚫")
        return

    todos_los_datos = []
    total_general = 0

    logging.info(f"Total de archivos PDF a procesar: {len(pdf_files)}")

    for file_path in pdf_files:
        logging.info(f"Procesando: {file_path} 📄")
        texto = extraer_texto_pdf(file_path)
        if not texto:
            logging.warning(f"No se pudo leer el texto del PDF: {os.path.basename(file_path)} ⚠️")
            continue
        datos = extraer_datos(texto)
        todos_los_datos.append(datos)

        logging.info("Datos extraídos: 📊")
        for k, v in datos.items():
            if k == "Servicios Detalle":
                logging.info(f"- {k}:")
                for s in v:
                    logging.info(f"   • {s.get('Descripción', 'N/A')} = ${s.get('Valor', 0):,.2f}")
            else:
                if isinstance(v, (int, float)) and k not in ["Días"]:
                    logging.info(f"- {k}: ${v:,.2f}" if v is not None else f"- {k}: N/A")
                else:
                    logging.info(f"- {k}: {v if v is not None else 'N/A'}")

        if datos.get("Total a Pagar") is not None:
            total_general += datos["Total a Pagar"]

    logging.info(f"Total general de todos los recibos: ${total_general:,.2f} 💰")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo_excel = f"informe_recibos_claro_{timestamp}.xlsx"
    
    # Intentar generar y subir el Excel
    try:
        exportar_excel_multiple(todos_los_datos, nombre_archivo_excel, total_general, FOLDER_ID_GOOGLE_DRIVE)
    except Exception as e:
        logging.error(f"Error al generar o subir el Excel '{nombre_archivo_excel}': {e} ❌")

    # Subir PDFs incluso si falla el Excel
    logging.info("Subiendo PDFs a Google Drive... 📤")
    for file_path in pdf_files:
        pdf_filename = os.path.basename(file_path)
        logging.info(f"Subiendo PDF: {pdf_filename}")
        try:
            uploaded_file_id = upload_file_to_drive(file_path, pdf_filename, FOLDER_ID_GOOGLE_DRIVE, is_file_path=True)
            if not uploaded_file_id:
                logging.error(f"Falló la subida de '{pdf_filename}' a Google Drive. ⚠️")
        except Exception as e:
            logging.error(f"Error al subir el PDF '{pdf_filename}' a Google Drive: {e} ❌")

    logging.info("Procesamiento y subida completados. Revisa tu Google Drive. ✅")

if __name__ == "__main__":
    main()
	
	
	

	